#!/usr/bin/env python3
"""Exporta db/chile.csv a un XLSX final con campos esenciales (incluye región y tipo de teléfono)."""
import csv, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import common as C

csv.field_size_limit(sys.maxsize)
DST = C.REPORTS / "chile-tiendas.xlsx"

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
BAD_EXT = (".gif",".png",".jpg",".jpeg",".svg",".webp",".css",".js",".ico")
def first_email(raw):
    for e in (raw or "").split(","):
        e=e.strip().lstrip(">").strip().lower(); e=re.sub(r"^u00[0-9a-f]{2}","",e)
        if EMAIL_RE.match(e) and not e.endswith(BAD_EXT): return e
    return ""
def fmt_phone(p):
    d=re.sub(r"\D","",p or "")
    if not d: return ""
    if d.startswith("56"): d=d[2:]
    if len(d)==9: return f"+56 {d[0]} {d[1:5]} {d[5:]}"
    if len(d)==8: return f"+56 {d[:4]} {d[4:]}"
    return "+56 "+d
def clean_addr(t,a):
    pref=f"{t} - "; return a[len(pref):] if a.startswith(pref) else a

def main():
    if not C.MASTER.exists():
        print("No existe db/chile.csv. Corré build_master.py primero."); return
    rows=list(csv.DictReader(open(C.MASTER,encoding="utf-8")))
    rows.sort(key=lambda x:(x.get("region",""),x.get("comuna",""),-int(x["review_count"] or 0)))
    headers=["Región","Comuna","Tienda","Categoría","Teléfono","Tipo","Email","★","Reseñas"]
    wb=Workbook(); ws=wb.active; ws.title="Tiendas Chile"
    hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(bold=True,color="FFFFFF")
    thin=Side(style="thin",color="D9D9D9"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.fill=hf; cell.font=hfont
        cell.alignment=Alignment(horizontal="center"); cell.border=bd
    for r,x in enumerate(rows,2):
        tel=fmt_phone(x["phone"])
        tt="Celular" if tel.startswith("+56 9") else "Fijo" if tel.startswith("+56 2") else ("Otro" if tel else "")
        vals=[x.get("region",""),x.get("comuna",""),x["title"],(x.get("category","") or "—").strip() or "—",
              tel or "—",tt,first_email(x.get("emails","")) or "—",
              round(float(x["review_rating"] or 0),1),int(x["review_count"] or 0)]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.border=bd
            cell.alignment=Alignment(horizontal="center" if c in (5,6,8,9) else "left")
            if r%2==0: cell.fill=PatternFill("solid",fgColor="F2F6FB")
    for c,w in enumerate([16,16,28,20,16,9,28,7,9],1):
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    wb.save(DST)
    print(f"OK -> {DST} ({len(rows)} negocios)")

if __name__ == "__main__":
    main()
